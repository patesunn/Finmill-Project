"""

Status of Project: In Development (Final touches)
% Completed: 95%
"""

#This program converts every user inputs to string and only converts to other data type when required

import tkinter as tk
import openpyxl as xl
import os
import datetime
import time
import math
from PIL import Image, ImageTk
import matplotlib.pyplot as plt

#some date and time properties str(datetime.datetime.now().strftime('%X'))
TODAY = datetime.datetime.now()
DATE = str(TODAY.strftime('%B') + " " + TODAY.strftime('%d') + ", " + TODAY.strftime("%Y"))
DAY = str(TODAY.strftime('%A'))
TIME = str(TODAY.strftime('%X'))

#some macros in program
LOGO_DISPLAY = 0
FULL_SCREEN = True
BACKGROUND_COLOR = "#397a1b"
HIGHLIGHT_COLOR = "orange"
FONT_COLOR = ""
FONT_TYPE = "Arial"

#Mathematical Parameters
FIN_HEIGHT_LOW = 9.02  
FIN_HEIGHT_HIGH = 9.08

#setting the paths for file saving
SAVE_PATH_1 = "D:\\Programming\\Python\\MCC- Internship\\Finmill Project" #local computer path
SAVE_PATH_2 = "D:\\Canadian Citizenship" #Network drive path
OPEN_PATH_FILE = "D:\\Programming\\Python\\MCC- Internship\\Finmill Project" #local computer path

class excelData:
	def __init__(self):
		"""
		Some class variables
		"""
		self.coreName = []
		self.coreNumber = []
		self.finNumber = []
		self.shape = []
		self.length = []
		self.tolerance = []
		self.convolutions = []
		self.finsPerInch = []
		self.finsPerPart = []

	
	def coreData(self):
		"""
		Function to extract data from excel file
		"""

		#changing working directory
		os.chdir(OPEN_PATH_FILE)

		#loads entire excel file into one variable
		initialData = xl.load_workbook("Condenser's fin specifications.xlsx")

		#loads sheet 1 into sheet1 variable
		sheet1 = initialData['Sheet1']

		#counts total entries in sheet1
		totalRows = sheet1.max_row

		#stores the data from excel files into list
		for i in range(totalRows):
			self.coreName.append(sheet1[chr(65)+str(i+2)].value)
			self.coreNumber.append(sheet1[chr(66)+str(i+2)].value)
			self.finNumber.append(sheet1[chr(67)+str(i+2)].value)
			self.shape.append(sheet1[chr(68)+str(i+2)].value)
			self.length.append(sheet1[chr(69)+str(i+2)].value)
			self.tolerance.append(sheet1[chr(70)+str(i+2)].value)
			self.convolutions.append(sheet1[chr(71)+str(i+2)].value)
			self.finsPerInch.append(sheet1[chr(72)+str(i+2)].value)
			self.finsPerPart.append(sheet1[chr(73)+str(i+2)].value)

		initialData.close()

		return self.coreName, self.coreNumber, self.finNumber, self.shape, self.length, self.tolerance, self.convolutions, self.finsPerInch, self.finsPerPart

class GUIinterface:

	def __init__(self, a, b, c, d, e, f, g, h, i):
		"""setting the window for GUI application"""		
		self.root=tk.Tk()
		self.root.configure(bg= BACKGROUND_COLOR) # 
		self.root.attributes('-fullscreen', FULL_SCREEN)

		"""Taking all the information from excelData class into variables (Bad way)"""
		self.coreName = a
		self.coreNumber = b
		self.finNumber = c
		self.shape = d
		self.length = e
		self.tolerance = f
		self.convolutions = g
		self.finsPerInch = h
		self.finsPerPart = i

		"""selected condenser from dropdown. The data is stored in this variable below"""
		self.select = str
		self.user_list=[]

#Basic setup details starts here including GUI display
	def mainLoop(self):
		"""
		Function for GUI to appear
		"""
		self.root.mainloop()

	def dateandtime(self):
		"""Date and Time (Maybe for later application)"""
		dateTimeLabel = tk.Label(text="{}; {}; {}".format(DAY, DATE, TIME), bd = 3)
		dateTimeLabel.config(font=(f"{FONT_TYPE}, 20"))
		dateTimeLabel.place(relx=0.250, rely=0.925, anchor='center')

	def companyLogo(self):
		"""Create a function for company's logo to appear at top"""

	def initialSetup(self):
		"""
		Main screen: the app starts from here
		"""
		self.root.title("Welcome to MCC's FinMill Database")
		
		label = tk.Label(text="EMC FinMill Database", fg='white', bg = BACKGROUND_COLOR)
		label.config(font=(FONT_TYPE, 72, 'bold'))
		
		#use label feature
		label.place(relx=0.500, rely=0.100, anchor="center")

#All the Buttons starts from here
	def setupOption(self):
		#creating first button
		button = tk.Button(self.root, text = "Setup Check", bd = '5', command=lambda:self.btnInputTrial(0), height='2', width='20', font=(FONT_TYPE, 30), state='active')
	
		#use place command to place the button at any place on screen
		button.place(relx=0.25, rely=0.45, anchor="center")

	def qCheckOption(self):
		#creating first button
		button = tk.Button(self.root, text = "Quality Check", bd = '5', command=lambda:self.btnInputTrial(1), height='2', width='20', font=(FONT_TYPE, 30), state='active')
	
		#use place command to place the button at any place on screen
		button.place(relx=0.75, rely=0.45, anchor="center")

	def genGraphs(self):
		#creating first button
		button = tk.Button(self.root, text = "Graphs", bd = '5', command=lambda:self.btnInputTrial(2), height='2', width='20', font=(FONT_TYPE, 30), state='active')
	
		#use place command to place the button at any place on screen
		button.place(relx= 0.5, rely=0.70, anchor="center")

	def exitButton(self):
		#creating exit button
		button = tk.Button(self.root, text = "Exit Program", command=lambda:self.btnInputTrial(4), font=(FONT_TYPE, 20), state='active')

		#placing "Exit Program" button for top right
		button.place(relx=0.98, rely=0.9, anchor='ne')

		#This function returns "False". This would help to exit the "While loop" and variable would be updated as well 
		return False

	def submitButton(self, state):
		#creating first button
		#The submit button for now would exit the application. Still have to figure out how to get values using .get() function multiple times. 
		# Error: _tkinter.TclError: invalid command name ".!entry"

		submitB = tk.Button(self.root, text = "Submit", bd = '5', height='1', width='10', font=(FONT_TYPE, 25), state=state, command=lambda:self.btnInputTrial(3))
	
		#use place command to place the button at any place on screen
		submitB.place(relx= 0.75, rely=0.950, anchor="center")

	def backButton(self):
		#creating a button to take back to main screen
		backB = tk.Button(self.root, text= "Back", bd = '5', height='1', width='10', font=(FONT_TYPE, 25), command=lambda:self.btnInputTrial(3))
		backB.place(relx=0.25, rely=0.950, anchor="center")

#Internally working functions
	def destroyApp(self):
		#this closes the program window and brings user back to windows screen
		self.root.destroy()

	def hideFrontScreen(self):
		#Hiding the front screen option as soon as option is been choosen
		for widget in self.root.winfo_children():
			widget.place_forget()

#Second screen functions starts here
	def setupLogs(self):

#SOME INTERNAL FUNCTIONS TO DO THE JOB
		#creating check button to see if user entered all required and correct data
		def checkButton():
		#if all the numbers are integers and/or floating, it would activate the submit button
			checkB = tk.Button(self.root, text = "Check", bd = '5', height='1', width='10', font=(FONT_TYPE, 25), state='active', command=entryCheck)
		#use place command to place the button at any place on screen
			checkB.place(relx= 0.50, rely=0.950, anchor="center")

		#creating a function that would check if all the entered values are within the specs
		def entryCheck():
			"""
			This function would check if the user has entered all numbers in text box. If not it would prompt them until a number is entered.
			"""
			#defining a variable to check if all users input are correct
			check_var = False

			a, b, c, d, e, f, g, h, i, j, k, l = menu.get(), coilSN.get("1.0","end-1c"), density_var.get(), length_var.get(), textBox.get("1.0","end-1c"), operatorText.get("1.0","end-1c"), scrapTotal.get("1.0","end-1c"), self.user_list[0].get(), self.user_list[1].get(), self.user_list[2].get(), self.user_list[3].get(), self.user_list[4].get()

			#creating a list grab all the values from user
			values_GUI = [a, b, c, d, e]
			height_user = [f, g, h, i, j, k, l]

			#validating if the entry is a number or alphanumeric
			for user in height_user:
				"""
				"""
				try:
					float(user)
					check_var = True

				except ValueError:
					check_var = False
					self.noValueCard()
					break

			if check_var == True:
				#creating an instance of class write and calling the function writeToFile() after that
				#def __init__(self, partNumber, category, coilnumber, employee, m0, m1, m2, m3, m4, density, length, scrap, notes):
				f = write(a, "Setup", b, f, h, i, j, k, l, c, d, g, e)
				f.writeToFile()

				del f

				#sleeps for 1.5 seconds to show off
				time.sleep(1.5)

				self.submitButton('active')

			del values_GUI, height_user

		#defining function for highlighting selected entry in Density
		def changeColorforDensity(s):
			"""
			Changing color and size based on user selection
			"""
			if s.lower() == 'yes':
				density_1.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				density_2.config(font=(FONT_TYPE, 28), fg='grey')

			elif s.lower() == 'no':
				density_2.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				density_1.config(font=(FONT_TYPE, 28), fg='grey')

		#defining function for highlighting selected entry in Length
		def changeColorforLength(s):
			"""
			Changing color and size based on user selection
			"""
			if s.lower() == 'yes':
				length_1.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				length_2.config(font=(FONT_TYPE, 28), fg='grey')

			elif s.lower() == 'no':
				length_2.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				length_1.config(font=(FONT_TYPE, 28), fg='grey')


# Screen's input parameters and labels starts from here, Finmill setup function only.
#-------------------------------------------------------------------------------------------------------------------------------------------
		#Check button to initiate checking if all the numbers entered are decimal 
		checkButton()

		#adding back button
		self.backButton()

		#Displaying the top label
		label = tk.Label(text="Finmill Setup Check", fg='orange', bg = BACKGROUND_COLOR)
		label.config(font=(FONT_TYPE, 56, 'bold'))
		label.place(relx=0.500, rely=0.075, anchor='center')

		#creating a label to name the dropdown menu 
		dropdownLabel = tk.Label(text="Condenser P/N: ")
		dropdownLabel.config(font=(FONT_TYPE, 18), height=2, bg = BACKGROUND_COLOR, fg= 'white')
		dropdownLabel.place(relx=0.025, rely=0.175)

		#creating dropdown/ Options menu
		#using * to extract values from the list
		menu = tk.StringVar(self.root)
		dropdown = tk.OptionMenu(self.root, menu, *self.coreNumber, command=self.printSelect)
		menu.set("Select condenser")
		dropdown.config(height = 2, font=("Arial", 14))
		dropdown.place(relx=0.160, rely=0.175)

		#creating a label for "Fin Stock Coil#" display\
		coilNumLabel = tk.Label(text="Fin Stock Coil# ")
		coilNumLabel.config(font=(FONT_TYPE, 18), height=2, bg = BACKGROUND_COLOR, fg='white')
		coilNumLabel.place(relx=0.325, rely=0.175)

		#Coil S/N: User input, alphanumeric
		coilSN = tk.Text(self.root, height = 1, width = 10, font=(FONT_TYPE, 36))
		coilSN.place(relx=0.445, rely=0.175)

		#operator's name and ID label
		operatorLabel = tk.Label(text="Operator's ID:")
		operatorLabel.config(font=(FONT_TYPE, 18), height=2, bg= BACKGROUND_COLOR, fg="white")
		operatorLabel.place(relx=0.70, rely=0.175)

		#operator's details textbox
		operatorText = tk.Text(self.root, height=1, width=10, font=(FONT_TYPE, 34))
		operatorText.place(relx=0.810, rely=0.175)

		#Label for Fin's P/N
		finPN = tk.Label(text="Fin's P/N: ")
		finPN.config(font=(FONT_TYPE, 18), height=2, bg=BACKGROUND_COLOR, fg='white')
		finPN.place(relx=0.325, rely=0.25)

		#Label for Condenser's P/N
		conPN = tk.Label(text="Condenser Name:")
		conPN.config(font=(FONT_TYPE, 18), height=2, bg=BACKGROUND_COLOR, fg='white')
		conPN.place(relx=0.025, rely=0.25)

		#Display label: "Parameters" & "Measurements"
		parameterLabel = tk.Label(text="Parameters", borderwidth=4, relief='raised')
		parameterLabel.config(font=(FONT_TYPE, 28, 'bold'), height=1, width= 20, bg= BACKGROUND_COLOR, fg="white")
		parameterLabel.place(relx=0.333, rely=0.375, anchor='e')

		measurementLabel = tk.Label(text="Measurements", borderwidth=4, relief='raised')
		measurementLabel.config(font=(FONT_TYPE, 28, 'bold'), height=1, width= 30, bg= BACKGROUND_COLOR, fg="white")
		measurementLabel.place(relx=0.667, rely=0.375, anchor='center')

		#creating the numbering sequence to just create an iteration. Nothing else
		numbersColumn = ["#1","#2","#3","#4","#5"]
		for i in range(5):
			numbersColumn[i] = tk.Label(text=f"#{i+1}")
			numbersColumn[i].config(font=(FONT_TYPE, 28), height=1, width= 3, bg= BACKGROUND_COLOR, fg="white")
			numbersColumn[i].place(relx=(0.3+0.115*(i+1)), rely=0.425)

		#creating a check box for length and density input from user
		density_var = tk.StringVar()
		length_var = tk.StringVar()

		density_1 = tk.Radiobutton(self.root, text='Yes', variable=density_var, value="Yes", command=lambda:changeColorforDensity("Yes"))
		density_1.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		density_1.place(relx=0.405, rely=0.575)

		density_2 = tk.Radiobutton(self.root, text='No', variable=density_var, value = "No", command=lambda:changeColorforDensity("No"))
		density_2.config(font=(FONT_TYPE, 28),bg=BACKGROUND_COLOR, height=1, width=3)
		density_2.place(relx=0.520, rely=0.575)

		length_1 = tk.Radiobutton(self.root, text='Yes', variable=length_var, value= "Yes", command=lambda:changeColorforLength("Yes"))
		length_1.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		length_1.place(relx=0.405, rely=0.675)

		length_2 = tk.Radiobutton(self.root, text='No', variable=length_var, value= "No", command=lambda:changeColorforLength("No"))
		length_2.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		length_2.place(relx=0.520, rely=0.675)

		#displaying the "Standard" Label for user's reference. This data would be copied from the base excel sheet, i.e. Condenser's fin specifications.xlsx
		standardLabel = tk.Label(text="Standard (mm)", borderwidth=2, relief='raised')
		standardLabel.config(font=(FONT_TYPE, 28), height=1, width=12, bg= 'blue', fg='white')
		standardLabel.place(relx=0.150, rely=0.425)

		#creating extry boxes for User Entry
		#total columns
		for y in range(1):
			#total rows in one column
			for x in range(5):
				my_list = tk.Entry(self.root, font=(FONT_TYPE, 32), width=4)
				my_list.place(relx=(0.2925+0.115*(x+1)), rely=0.485+(0.1*y))
				#this is just a raw data. Use .get() function to extract the values from the textbox
				self.user_list.append(my_list)

		#getting only last 5 entries of the existing list, self.user_list
		self.user_list = self.user_list[-5::]

		parameters_list = ['Height', 'Density', 'Length']
		for x in range(3):
			parameters_list[x] = tk.Label(text=parameters_list[x])
			parameters_list[x].config(font=(FONT_TYPE, 30), height=1, width=8, bg= BACKGROUND_COLOR, fg='white')
			parameters_list[x].place(relx=0.025, rely=0.485+(0.1*x))

		#scrapped fins entry 
		scrapLabel = tk.Label(text='Fins scrapped: ')
		scrapLabel.config(font=(FONT_TYPE, 18), bg=BACKGROUND_COLOR, fg='white')
		scrapLabel.place(relx=0.7, rely=0.2625)

		#entry label for scrapped label
		scrapTotal = tk.Text(self.root, height = 1, width = 10, font=(FONT_TYPE, 34))
		scrapTotal.place(relx=0.81, rely=0.25)

		#Label for any additional comments
		textLabel = tk.Label(text="Comments")
		textLabel.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, fg='white')
		textLabel.place(relx=0.025, rely=0.80)

		#entry box for any user comments
		textBox = tk.Text(self.root, height=1, width=50, font=(FONT_TYPE, 34))
		textBox.place(relx=0.155, rely=0.80)
#---------------------------------------------------------------------------------------------------------------------------------------
	#This function would allow to store the data into a variable for later use, i.e. storing data into excel file
	#It stores the part number of the condenser that user has selected in this menu
	def printSelect(self, name):
		"""
		This variable stores the entry which is selected by the user
		"""
		self.name = name
		take_index = self.coreNumber.index(self.name)

		#printing the parameters of the selected condenser on the screen
		specs_condenser = ["9.02 - 9.08", self.finsPerInch[take_index], self.length[take_index], (str(self.finNumber[take_index]) + " (" + str(self.shape[take_index]) + ")"), str(self.coreName[take_index])]

		for x in range(3):
			specs_condenser[x] = tk.Label(text=specs_condenser[x])
			specs_condenser[x].config(font=(FONT_TYPE, 28), height=1, width=8, bg= BACKGROUND_COLOR, fg='white')
			specs_condenser[x].place(relx=0.175, rely=0.485+(0.1*x))
		
		#creating a label for Fin's P/N display
		finPNdisplay = tk.Label(text=specs_condenser[3])
		finPNdisplay.config(font=(FONT_TYPE,20), height=1, width=20, bg= BACKGROUND_COLOR, fg= 'white')
		finPNdisplay.place(relx=0.4, rely=0.2585)

		#creating a label for Condenser's P/N display
		conPNdisplay = tk.Label(text=specs_condenser[4])
		conPNdisplay.config(font=(FONT_TYPE,20), height=1, width=15, bg= BACKGROUND_COLOR, fg= 'white')
		conPNdisplay.place(relx=0.155, rely=0.2585)
		

	def qualityLogs(self):
		#This function would be identical to Setup Logs. 
		#Only difference would be, the user cannot change the product name since the machine is already setup
		#We can use same code, but it would get bit complicated.

		#creating check button to see if user entered all required and correct data
		def checkButton():
		#if all the numbers are integers and/or floating, it would activate the submit button
			checkB = tk.Button(self.root, text = "Check", bd = '5', height='1', width='15', font=(FONT_TYPE, 25), state='active', command=entryCheck)
		#use place command to place the button at any place on screen
			checkB.place(relx= 0.50, rely=0.950, anchor="center")


		#creating a function that would check if all the entered values are within the specs
		def entryCheck():
			"""
			This function would check if the user has entered all numbers in text box. If not it would prompt them until a number is entered.
			"""
			#defining a variable to check if all users input are correct (and assuming that they are correct)
			check_var = True

			#add conditions of noValueCard() and rightValueCard()		
			while check_var: 

				a, b, c, d, e, f, g, h, i, j, k, l = conPNDisplay["text"], coilNumLabel["text"], density_var.get(), length_var.get(), textBox.get("1.0","end-1c"), operatorText.get("1.0","end-1c"), scrapTotal.get("1.0","end-1c"), self.user_list[0].get(), self.user_list[1].get(), self.user_list[2].get(), self.user_list[3].get(), self.user_list[4].get()

				#creating a list grab all the values from user
				values_GUI = [a, b, c, d, e]
				height_user = [f, g, h, i, j, k, l]

				#validating if the entry is a number or alphanumeric
				for user in height_user:
					"""
					"""
					try:
						float(user)
						check_var = True

					except ValueError:
						check_var = False
						self.noValueCard()
						break

				#creating an instance of class write and calling the function writeToFile() after that
				#def __init__(self, partNumber, category, coilnumber, employee, m0, m1, m2, m3, m4, density, length, scrap, notes):
				f = write(a, "Quality", b, f, h, i, j, k, l, c, d, g, e)
				f.writeToFile()

				del f

				#sleeps for 1.5 seconds to show off
				time.sleep(1.5)

				del values_GUI, height_user

				check_var = rightValueCard()

				#It checks if all the entries are within the specs, it would terminate the loop
				if check_var == True:
					self.submitButton('active')
					break
				else:
					pass

		#Adding a check button for entry
		checkButton()

		#adding back button
		self.backButton()

		def getLastPN():
			"""
			* This function would read "Finmill Data Logs.xlsx" file and would search for "Setup" in column F
			* Once found it would store the row number. It would be a loop that searches for the keyword and record the last entry
			* After loop completes the operation, it would get the associated part number with that row
			* This function would return the index of the part number from self.coreNumber
			* The parameters for that part number would be displayed on screen. This serves as an reference for user.
			"""
			#creating a variable to store running position
			index_pos = int

			#creating an instance of workbook
			f = xl.load_workbook("Finmill Logged Data.xlsx")
			#creating an instance of worksheet
			sheet1 = f.worksheets[0]
			#Getting the last row from sheet1
			last_row = sheet1.max_row

			#looping thru entire worksheet to find last setup entry
			for i in range(last_row):
				if (sheet1[chr(70)+str(i+1)].value).strip() == 'Setup':
					last_row = i+1

			return self.coreNumber.index(sheet1[chr(68)+str(last_row)].value), sheet1[chr(69)+str(last_row)].value

		#defining function for highlighting selected entry in Density
		def changeColorforDensity(s):
			"""
			Changing color and size based on user selection
			"""
			if s.lower() == 'yes':
				density_1.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				density_2.config(font=(FONT_TYPE, 28), fg='grey')

			elif s.lower() == 'no':
				density_2.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				density_1.config(font=(FONT_TYPE, 28), fg='grey')

		#defining function for highlighting selected entry in Length
		def changeColorforLength(s):
			"""
			Changing color and size based on user selection
			"""
			if s.lower() == 'yes':
				length_1.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				length_2.config(font=(FONT_TYPE, 28), fg='grey')

			elif s.lower() == 'no':
				length_2.config(font=(FONT_TYPE, 28, 'bold'), fg=HIGHLIGHT_COLOR)
				length_1.config(font=(FONT_TYPE, 28), fg='grey')


		#Defining a variable for storing the last row value for which "Setup" exist	and getting last running coil
		position, coil = getLastPN()

		#Displaying the top label
		label = tk.Label(text="Finmill Quality Check", fg='orange', bg = BACKGROUND_COLOR)
		label.config(font=(FONT_TYPE, 56, 'bold'))
		label.place(relx=0.500, rely=0.075, anchor='center')

		#operator's name and ID label
		operatorLabel = tk.Label(text="Operator's ID:")
		operatorLabel.config(font=(FONT_TYPE, 18), height=2, bg= BACKGROUND_COLOR, fg="white")
		operatorLabel.place(relx=0.70, rely=0.175)

		#creating a label to name the dropdown menu 
		dropdownLabel = tk.Label(text="Condenser Name: ")
		dropdownLabel.config(font=(FONT_TYPE, 18), height=2, bg = BACKGROUND_COLOR, fg= 'white')
		dropdownLabel.place(relx=0.025, rely=0.175)

		#displaying condenser name based on last setup
		condName = tk.Label(text=self.coreName[position])
		condName.config(font=(FONT_TYPE, 18), height=2, bg= BACKGROUND_COLOR, fg= 'white')
		condName.place(relx= 0.160, rely=0.175)

		#creating a label for "Fin Stock Coil#" display\
		coilNumLabel = tk.Label(text="Fin Stock Coil# ")
		coilNumLabel.config(font=(FONT_TYPE, 18), height=2, bg = BACKGROUND_COLOR, fg='white')
		coilNumLabel.place(relx=0.325, rely=0.175)

		#Displaying the fin stock #
		coilNumLabel = tk.Label(text=coil)
		coilNumLabel.config(font=(FONT_TYPE, 18), height=2, bg = BACKGROUND_COLOR, fg='white')
		coilNumLabel.place(relx=0.445, rely=0.175)

		#Label for Fin's P/N
		finPN = tk.Label(text="Fin's P/N: ")
		finPN.config(font=(FONT_TYPE, 18), height=2, bg=BACKGROUND_COLOR, fg='white')
		finPN.place(relx=0.325, rely=0.25)

		#displaying fin P/N
		finDisplay = tk.Label(text=self.finNumber[position])
		finDisplay.config(font=(FONT_TYPE,18), height=2, bg=BACKGROUND_COLOR, fg='white')
		finDisplay.place(relx=0.4, rely=0.25)

		#Label for Condenser's P/N
		conPN = tk.Label(text="Condenser P/N:")
		conPN.config(font=(FONT_TYPE, 18), height=2, bg=BACKGROUND_COLOR, fg='white')
		conPN.place(relx=0.025, rely=0.25)

		#displaying condenser P/N
		conPNDisplay = tk.Label(text=self.coreNumber[position])
		conPNDisplay.config(font=(FONT_TYPE,18), height=2, bg=BACKGROUND_COLOR, fg='white')
		conPNDisplay.place(relx=0.145, rely=0.25)

		#Display label: "Parameters" & "Measurements"
		parameterLabel = tk.Label(text="Parameters", borderwidth=4, relief='raised')
		parameterLabel.config(font=(FONT_TYPE, 28, 'bold'), height=1, width= 20, bg= BACKGROUND_COLOR, fg="white")
		parameterLabel.place(relx=0.333, rely=0.375, anchor='e')

		measurementLabel = tk.Label(text="Measurements", borderwidth=4, relief='raised')
		measurementLabel.config(font=(FONT_TYPE, 28, 'bold'), height=1, width= 30, bg= BACKGROUND_COLOR, fg="white")
		measurementLabel.place(relx=0.667, rely=0.375, anchor='center')

		#displaying the "Standard" Label for user's reference. This data would be copied from the base excel sheet, i.e. Condenser's fin specifications.xlsx
		standardLabel = tk.Label(text="Standard (mm)", borderwidth=2, relief='raised')
		standardLabel.config(font=(FONT_TYPE, 28), height=1, width=12, bg= 'blue', fg='white')
		standardLabel.place(relx=0.150, rely=0.425)

		parameters_list = ['Height', 'Density', 'Length']
		for x in range(3):
			parameters_list[x] = tk.Label(text=parameters_list[x])
			parameters_list[x].config(font=(FONT_TYPE, 30), height=1, width=8, bg= BACKGROUND_COLOR, fg='white')
			parameters_list[x].place(relx=0.025, rely=0.485+(0.1*x))

		#printing the parameters of the selected condenser on the screen
		specs_condenser = ["9.02 - 9.08", self.finsPerInch[position], self.length[position]]

		for x in range(3):
			specs_condenser[x] = tk.Label(text=specs_condenser[x])
			specs_condenser[x].config(font=(FONT_TYPE, 28), height=1, width=8, bg= BACKGROUND_COLOR, fg='white')
			specs_condenser[x].place(relx=0.175, rely=0.485+(0.1*x))

		#creating the numbering sequence to just create an iteration. Nothing else
		numbersColumn = ["#1","#2","#3","#4","#5"]
		for i in range(5):
			numbersColumn[i] = tk.Label(text=f"#{i+1}")
			numbersColumn[i].config(font=(FONT_TYPE, 28), height=1, width= 3, bg= BACKGROUND_COLOR, fg="white")
			numbersColumn[i].place(relx=(0.3+0.115*(i+1)), rely=0.425)

		#scrapped fins entry 
		scrapLabel = tk.Label(text='Fins scrapped: ')
		scrapLabel.config(font=(FONT_TYPE, 18), bg=BACKGROUND_COLOR, fg='white')
		scrapLabel.place(relx=0.70, rely=0.2625)

		#Label for any additional comments
		textLabel = tk.Label(text="Comments")
		textLabel.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, fg='white')
		textLabel.place(relx=0.025, rely=0.80)


	#Things that require user input in Quality Check
		#operator's details textbox
		operatorText = tk.Text(self.root, height=1, width=10, font=(FONT_TYPE, 34))
		operatorText.place(relx=0.810, rely=0.175)

		#entry label for scrapped label
		scrapTotal = tk.Text(self.root, height = 1, width = 10, font=(FONT_TYPE, 34))
		scrapTotal.place(relx=0.810, rely=0.25)

		#creating extry boxes for User Entry
		#total columns
		for y in range(1):
			#total rows in one column
			for x in range(5):
				my_list = tk.Entry(self.root, font=(FONT_TYPE, 32), width=4)
				my_list.place(relx=(0.2925+0.115*(x+1)), rely=0.485+(0.1*y))
				#this is just a raw data. Use .get() function to extract the values from the textbox
				self.user_list.append(my_list)

		#getting only last 5 entries of the existing list, self.user_list
		self.user_list = self.user_list[-5::]

		#creating a check box for length and density input from user
		density_var = tk.StringVar()
		length_var = tk.StringVar()

		density_1 = tk.Radiobutton(self.root, text='Yes', variable=density_var, value="Yes", command=lambda:changeColorforDensity("Yes"))
		density_1.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		density_1.place(relx=0.405, rely=0.575)

		density_2 = tk.Radiobutton(self.root, text='No', variable=density_var, value = "No", command=lambda:changeColorforDensity("No"))
		density_2.config(font=(FONT_TYPE, 28),bg=BACKGROUND_COLOR, height=1, width=3)
		density_2.place(relx=0.520, rely=0.575)

		length_1 = tk.Radiobutton(self.root, text='Yes', variable=length_var, value= "Yes", command=lambda:changeColorforLength("Yes"))
		length_1.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		length_1.place(relx=0.405, rely=0.675)

		length_2 = tk.Radiobutton(self.root, text='No', variable=length_var, value= "No", command=lambda:changeColorforLength("No"))
		length_2.config(font=(FONT_TYPE, 28), bg=BACKGROUND_COLOR, height=1, width=3)
		length_2.place(relx=0.520, rely=0.675)

		#entry box for any user comments
		textBox = tk.Text(self.root, height=1, width=50, font=(FONT_TYPE, 34))
		textBox.place(relx=0.155, rely=0.80)
		
		def rightValueCard():
			"""
			- This is an internal function for quality check only
			- It would be only called if all the entries are been filled by the user
			- It grabs the data from above mentioned data types and compares them to the standard values
			- If all the values are within the specs, it would play along and save the data
			- If the values are off, it would notify the user and sstill save the data
			- Returns True at end if all entries are correct
			- Returns False if one of the entries are incorrect
			"""
			# Getting all the data first, i.e. user entry only and only qualitative
			# Height is already been stored in self.user_list

			# Creating a running variable to check dimension status
			dimension = True

			def outOfSpecsFunc():
				#creates a new pop up box 
				message_box = tk.Tk()

				def deletePop():
					message_box.destroy()

				message_box.geometry("750x200")
				message_box.title("Out of specifications!!!")

				#places the new dialog box in almost center of screen
				message_box.eval('tk::PlaceWindow . center')

				message_pop = tk.Label(message_box, text="You have entered an out of specifications value(s) in one or more fields. \n\nPlease check and/or correct your entry!!!", font=(FONT_TYPE, 16))
				message_pop.place(relx=0.5, rely=0.25, anchor='center')

				exit_button = tk.Button(message_box, text='OK', font=(FONT_TYPE,20), height=1, width=10, bd=3, command=deletePop)
				exit_button.place(relx=0.5, rely=0.75, anchor='center')

				message_box.mainloop()

			for numbers in self.user_list:
				#Since the fin height is always constant, the macros are defined at top
				#Tried using the math.isclose() function but it doesnt regard for lower limit
				if math.isclose(9.05, float(numbers.get()), abs_tol=0.03)==True:
					dimension = True
				else:
					dimension = False
					break

			#checks for height and density entry as well
			if density_var.get().lower() == "no" or length_var.get().lower() == "no" or dimension == False:
				#once all the data entries are checked, it would call the out of spec window if any entry is out of specifications
				outOfSpecsFunc()

			else: pass

			#The error box is only displayed when any of the dimension is out of specs
			if dimension == True:
				return True
			elif dimension == False: 
				return False
			else:
				pass

#Decision functions starts here
	def btnInputTrial(self, args):
		if args == 0:
			#As user clicks on setup button, all the other options disappears. The exit button disappears too since user won't be able to close the app in middle.
			self.hideFrontScreen()
			self.setupLogs()

		elif args == 1:
			#As user clicks on quality button, all the other options disappears. The exit button disappears too since user won't be able to close the app in middle.
			self.hideFrontScreen()
			self.qualityLogs()

		elif args == 2:
			#Grabbing the last active part from excel file
			#creating an instance of workbook
			f = xl.load_workbook("Finmill Logged Data.xlsx")

			#creating an instance of worksheet
			sheet1 = f.worksheets[0]
			#getting maximum rows in spreadsheet
			f_maxrows = sheet1.max_row

			last_PN = sheet1[chr(68)+str(f_maxrows)].value

			#It takes in one argument
			g = Graphs(last_PN)
			g.generateGraphs()

			#deleting the local variables are use
			del g, f
	
		elif args == 3:
			self.hideFrontScreen()
			self.initialSetup()
			self.setupOption()
			self.qCheckOption()
			self.genGraphs()
			self.exitButton()

		else:
			self.destroyApp()

#Function to display an error message if the user has entered incorrect type of value
	def noValueCard(self):
		#this creates a small dialog box
		error_box = tk.Tk()
		error_box.geometry("750x200")
		error_box.title('NO Value!!!')

		#places the new dialog box in almost center of screen
		error_box.eval('tk::PlaceWindow . center')

		#function to delete pop window
		def deletePop():
			error_box.destroy()

		message_pop = tk.Label(error_box, text="NO value(s) in one or more fields. \n\nPlease fill all required entries!!!", font=(FONT_TYPE, 16))
		message_pop.place(relx=0.5, rely=0.25, anchor='center')

		exit_button = tk.Button(error_box, text='OK', font=(FONT_TYPE,20), height=1, width=10, bd=3, command=deletePop)
		exit_button.place(relx=0.5, rely=0.75, anchor='center')

		error_box.mainloop()

class write():
	def __init__(self, partNumber, category, coilnumber, employee, m0, m1, m2, m3, m4, density, length, scrap, notes):
		"""
		Stores all the names and part numbers for all the condensers available from base excel file
		write("T41-0017-00", "Setup", "218A", 2190, 9.05, 9.07, 9.08, 9.05, 9.05, "Yes", "Yes", 20, "You got this!!!")
		"""
		self.partNumber = partNumber
		self.category = category
		self.coilnumber = coilnumber
		self.employee = employee
		self.m0 = m0
		self.m1 = m1
		self.m2 = m2
		self.m3 = m3
		self.m4 = m4
		self.density = density
		self.length = length
		self.scrap = scrap
		self.notes = notes

		#preparing for easily writing to file
		self.allInOne = [DATE, str(datetime.datetime.now().strftime('%X')), int(self.employee), self.partNumber, self.coilnumber, self.category, float(self.m0), float(self.m1), float(self.m2), float(self.m3), float(self.m4), self.density, self.length, int(self.scrap), self.notes]

	def writeToFile(self):
		"""
		Function to write data to particular file.		
		Types of data filing format.
		Option 1: 2 Files (Setup and Quality), multiple sheets, one per condenser
		Option 2: 1 File per condenser, two sheets, i.e. setup and quality
		Option 3: 1 File per condenser, 1 sheet, one column for quality or setup

		Decided: 1 File, all logs into it; Furnace type data logging
		"""
		#creating an instance of workbook
		f = xl.load_workbook("Finmill Logged Data.xlsx")
		#creating an instance of worksheet
		sheet1 = f.worksheets[0]
		#getting maximum rows in spreadsheet
		f_maxrows = sheet1.max_row

		for i in range(len(self.allInOne)):
			"""
			Writing to each cell. Correct for positioning. 
			"""
			sheet1[chr(65+i)+str(f_maxrows+1)] = self.allInOne[i]

		os.chdir(SAVE_PATH_1)
		f.save("Finmill Logged Data.xlsx")

		os.chdir(SAVE_PATH_2)
		f.save("Finmill Logged Data.xlsx")

		os.chdir(SAVE_PATH_1)

		del f, sheet1, f_maxrows

class Graphs():
	"""
	-This class would contain all the data members and methods for generating a Control chart using matplotlib library
	-Based on the last "Setup" input from the user, this class's function would collect some data and graph it accordingly
	"""

	def __init__(self, partnumber):
		"""
		Defining some important parameters
		"""
		self.partnumber = partnumber


	def generateGraphs(self):
		"""
		"""
		#opening up the excel file to grab the data
		f = xl.load_workbook("Finmill Logged Data.xlsx")
		#setting up sheet 0 as active sheet
		sheet1 = f.worksheets[0]
		#getting maximum rows in spreadsheet
		f_maxrows = sheet1.max_row

		#creating an empty list to store all the values of y
		y_values = []

		for i in range(f_maxrows):
			"""
			The only thing edited
			"""
			if sheet1[chr(68)+str(i+1)].value == self.partnumber:
					y_values.append([sheet1[chr(71)+str(i+1)].value, sheet1[chr(72)+str(i+1)].value, sheet1[chr(73)+str(i+1)].value, sheet1[chr(74)+str(i+1)].value, sheet1[chr(75)+str(i+1)].value])

		#retains last 50 entries cummulative, setup and quality check		
		if len(y_values) >= 50:
			y_values = y_values[-50::]
		else:
			pass

		#max allowed values to display
		x_values = [x for x in range(1,11)]

		#Matplotlib library functions
		#Labeling the axis and chart
		plt.suptitle("Data points for {}".format(self.partnumber))
		plt.xlabel("Last 10 Entries (Date and Time)")
		plt.ylabel("Fin height (mm)")

		#setting the limits for y axis
		plt.ylim((8.995,9.11))

		#adding upper limit and lower limit lines
		x_axis = [x for x in range(1,11)]
		y_lower = [9.02, 9.02, 9.02,9.02,9.02,9.02,9.02,9.02,9.02,9.02]
		y_upper = [9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08]

		plt.plot(x_axis, y_lower, linestyle='dashed', color="red")
		plt.plot(x_axis, y_upper, linestyle='dashed', color="red")

		#Adding different types of markers for each set
		markers = ["o","v", "^", "s", "D", "o","v", "^", "s", "D"]

		for x,y,z in zip(x_values, y_values, markers):
			plt.scatter([x]*len(y), y, marker=z)

		plt.show()

		#deleting all variables
		del f, sheet1, f_maxrows, y_values, x_values, x_axis, y_lower, y_upper

def main():
	"""
	Reading data from the excel file
	"""
	#creating an instance of excelData class
	xl_file = excelData()
	
	#storing the data from file into variable for future use
	A, B, C, D, E, F, G, H, I = xl_file.coreData()

	#creating an instance for GUIinterface class
	gui = GUIinterface(A, B, C, D, E, F, G, H, I)

	gui.initialSetup()
	gui.setupOption()
	gui.qCheckOption()
	gui.genGraphs()

	gui.mainLoop()
			
	# Deleting the class variables before end of program	
	del xl_file
	del gui

main()