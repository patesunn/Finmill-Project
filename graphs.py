import matplotlib.pyplot as plt
import tkinter as tk
import openpyxl as xl

class graphs():
	def __init__(self, window, bg_color, fg_color):
		"""
		Add some data types in here
		"""
		#Takes in the window name from main file
		self.root = window
		self.root.configure(bg= "#ebb434")
		self.root.attributes('-fullscreen', True)
		self.bg_color = bg_color
		self.fg_color = fg_color

	def generateGraphs(self, condenserName):
		"""
		This function would take Condenser P/N from userWindow() function and would display the graph
		- Opens the file and prints last 10 entries
		"""
		#opening up the excel file to grab the data
		f = xl.load_workbook("Finmill Logged Data.xlsx")
		#setting up sheet 0 as active sheet
		sheet1 = f.worksheets[0]
		#getting maximum rows in spreadsheet
		f_maxrows = sheet1.max_row

		#creating an empty list to store all the values of y
		y_values = []

		for i in range(f_maxrows):
			"""
			The only thing edited
			"""
			if sheet1[chr(68)+str(i+1)].value == condenserName:
					y_values.append([sheet1[chr(71)+str(i+1)].value, sheet1[chr(72)+str(i+1)].value, sheet1[chr(73)+str(i+1)].value, sheet1[chr(74)+str(i+1)].value, sheet1[chr(75)+str(i+1)].value])

		#retains last 50 entries cummulative, setup and quality check		
		if len(y_values) >= 50:
			y_values = y_values[-50::]
		else:
			pass

		#max allowed values to display
		x_values = [x for x in range(1,11)]

		#Matplotlib library functions
		#Labeling the axis and chart
		plt.suptitle("Data points for {}".format(condenserName))
		plt.xlabel("Last 10 Entries (Date and Time)")
		plt.ylabel("Fin height (mm)")

		#setting the limits for y axis
		plt.ylim((8.995,9.11))

		#adding upper limit and lower limit lines
		x_axis = [x for x in range(1,11)]
		y_lower = [9.02, 9.02, 9.02,9.02,9.02,9.02,9.02,9.02,9.02,9.02]
		y_upper = [9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08, 9.08]

		plt.plot(x_axis, y_lower, linestyle='dashed', color="red")
		plt.plot(x_axis, y_upper, linestyle='dashed', color="red")

		#Adding different types of markers for each set
		markers = ["o","v", "^", "s", "D", "o","v", "^", "s", "D"]

		for x,y,z in zip(x_values, y_values, markers):
			plt.scatter([x]*len(y), y, marker=z)

		plt.show()

		#deleting all variables
		del f, sheet1, f_maxrows, y_values, x_values, x_axis, y_lower, y_upper


	def userWindow(self):
		"""
		"""
		f = xl.load_workbook("Condenser's fin specifications.xlsx")
	
		condensers = []

		for i in range(1,f.worksheets[0].max_row):
			"""
			"""
			condensers.append(f.worksheets[0][chr(66)+str(i+1)].value)

		del f

		title_lbl = tk.Label(text="Generate Graphs", fg=self.fg_color, bg=self.bg_color)
		title_lbl.config(font=("Arial", 72, 'bold'))
		title_lbl.place(relx=0.5, rely=0.10, anchor='center')

		#Label for "Select Condenser: "
		condenser_label = tk.Label(text="Select Condenser: ", fg= self.fg_color, bg= self.bg_color)
		condenser_label.config(font=("Arial", 28, 'bold'))
		condenser_label.place(relx=0.5, rely=0.4, anchor='e')

		#dropdown for selecting condenser
		conVar = tk.StringVar()
		cond_dropdown = tk.OptionMenu(self.root, conVar, *condensers)
		cond_dropdown.config(font=("Arial", 28, 'bold'), height=1, width=15)
		cond_dropdown.place(relx=0.5, rely=0.4, anchor='w')

		#Generate graph button
		graph_button = tk.Button(self.root, text="Generate", bd = '5', height='1', width='10', font=("Arial", 30), state='active', command=lambda:self.generateGraphs(conVar.get()))
		graph_button.place(relx=0.5, rely=0.6, anchor='center')

		#this statement would be deleted once we execute main file
		self.root.mainloop()

c = graphs(tk.Tk(), "#ebb434", "white")
c.userWindow()